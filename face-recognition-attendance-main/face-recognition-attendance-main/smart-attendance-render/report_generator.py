"""Automatic attendance report generation — Excel output with per-student
attendance %, using pandas + openpyxl."""
import os
from datetime import date
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import Config
from models import get_attendance_between, get_all_students


def generate_report(start_date: date, end_date: date, class_id=None) -> str:
    """Builds a student x date attendance matrix + summary sheet, saves as .xlsx,
    and returns the file path."""
    records = get_attendance_between(start_date, end_date, class_id)
    students = get_all_students(class_id)

    if not students:
        raise ValueError("No students found for the given class.")

    # Build lookup: (roll_number, date) -> status
    present_lookup = {}
    for r in records:
        key = (r["roll_number"], r["session_date"])
        present_lookup[key] = r["status"]

    session_dates = sorted({r["session_date"] for r in records})
    if not session_dates:
        session_dates = []

    rows = []
    for s in students:
        row = {"Roll No": s["roll_number"], "Name": s["name"]}
        present_count = 0
        for d in session_dates:
            status = present_lookup.get((s["roll_number"], d))
            mark = "P" if status == "present" else ("L" if status == "late" else "A")
            if mark in ("P", "L"):
                present_count += 1
            row[d.strftime("%Y-%m-%d")] = mark
        total_sessions = len(session_dates)
        row["Sessions Held"] = total_sessions
        row["Present"] = present_count
        row["Attendance %"] = round(100 * present_count / total_sessions, 1) if total_sessions else 0.0
        rows.append(row)

    df = pd.DataFrame(rows)

    os.makedirs(Config.REPORTS_FOLDER, exist_ok=True)
    filename = f"attendance_report_{start_date}_{end_date}.xlsx"
    filepath = os.path.join(Config.REPORTS_FOLDER, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Attendance", index=False)
        _style_worksheet(writer.sheets["Attendance"], df)

    return filepath


def _style_worksheet(ws, df):
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(col_name)) + 2)

    # Conditional highlight: attendance % below 75 in red-ish
    if "Attendance %" in df.columns:
        pct_col_idx = list(df.columns).index("Attendance %") + 1
        low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row_idx, val in enumerate(df["Attendance %"], start=2):
            if val < 75:
                ws.cell(row=row_idx, column=pct_col_idx).fill = low_fill


def generate_daily_summary_csv(session_date: date, class_id=None) -> str:
    """Quick CSV of who was present/absent for a single day."""
    records = get_attendance_between(session_date, session_date, class_id)
    students = get_all_students(class_id)
    present_rolls = {r["roll_number"] for r in records if r["status"] in ("present", "late")}

    rows = [
        {"Roll No": s["roll_number"], "Name": s["name"],
         "Status": "Present" if s["roll_number"] in present_rolls else "Absent"}
        for s in students
    ]
    df = pd.DataFrame(rows)
    os.makedirs(Config.REPORTS_FOLDER, exist_ok=True)
    filepath = os.path.join(Config.REPORTS_FOLDER, f"daily_{session_date}.csv")
    df.to_csv(filepath, index=False)
    return filepath
